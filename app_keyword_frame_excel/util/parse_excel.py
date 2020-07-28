import openpyxl
from openpyxl.styles import Border, Side, Font, colors
import time


from config.var_config import *
from util.log import *


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.test_case_file = None
        # set font color
        self.font = Font(color=None)
        # color and RGB dict
        self.RGB_dict = {'red': colors.RED, 'green': colors.GREEN, 'blue': colors.BLUE}

    def load_workbook(self, test_case_file):
        """
        load excel file, get the workbook object
        :param test_case_file: abs path
        :return: workbook object
        """
        try:
            self.workbook = openpyxl.load_workbook(test_case_file)
        except Exception as e:
            raise e
        # use in saving file
        self.test_case_file = test_case_file
        return self.workbook

    def get_sheet_by_name(self, sheet_name):
        # get sheet object by name
        try:
            sheet_obj = self.workbook[sheet_name]
            return sheet_obj
        except Exception as e:
            raise e

    def get_sheet_by_index(self, sheet_index):
        # get sheet object by index
        try:
            # sheet_index starts from 0
            sheet_name = self.workbook.sheetnames[sheet_index]
            sheet_obj = self.workbook[sheet_name]
            return sheet_obj
        except Exception as e:
            info("索引号为%d的sheet表不存在" % sheet_index)  # 后面转成log
            raise e

    def get_start_row(self, sheet_obj):
        # get the start row number
        try:
            return sheet_obj.min_row
        except Exception as e:
            raise e

    def get_end_row(self, sheet_obj):
        # get the end row number
        try:
            return sheet_obj.max_row
        except Exception as e:
            raise e

    def get_start_column(self, sheet_obj):
        # get the start column number
        try:
            return sheet_obj.min_column
        except Exception as e:
            raise e

    def get_end_column(self, sheet_obj):
        # get the end column number
        try:
            return sheet_obj.max_column
        except Exception as e:
            raise e

    def get_row_obj(self, sheet_obj, row_no):
        """
        get certain row object
        list starts with 0， and excel starts with 1， so need to -1
        :param row_no: int, row index, starts with 1, 1 means the first row !!!
        :return: row object, type: tuple
        """
        try:
            rows = []
            # sheet_obj.iter_rows(): <generator object Worksheet._cells_by_row at 0x000001C6623AB678>
            for row in sheet_obj.iter_rows():
                rows.append(row)
            return rows[row_no - 1]
        except Exception as e:
            raise e

    def get_column_obj(self, sheet_obj, col_no):
        try:
            cols = []
            for col in sheet_obj.iter_cols():
                cols.append(col)
            return cols[col_no - 1]
        except Exception as e:
            raise e


    def get_cell_value(self, sheet_obj, row_no = None, col_no = None):
        """
        get cell value
        :param row_no: int, row index, starts with 1！ means the first row
        :param col_no: int, column index, starts with 1！ means the first column
        :return: cell content
        """
        if row_no and col_no:
            try:
                return sheet_obj.cell(row=row_no, column=col_no).value
            except Exception as e:
                raise e
        else:
            raise Exception("get cell value: insufficient coordinates of cell")

    def get_cell_obj(self, sheet_obj, row_no=None, col_no=None):
        if row_no and col_no:
            try:
                return sheet_obj.cell(row=row_no, column=col_no)
            except Exception as e:
                raise e
        else:
            raise Exception("get cell obj: insufficient coordinates of cell")

    def write_cell(self, sheet_obj, row_no=None, col_no=None, content=None, style=None):
        if row_no and col_no:
            try:
                sheet_obj.cell(row=row_no, column=col_no).value = content
                if style and style in self.RGB_dict:
                    sheet_obj.cell(row=row_no, column=col_no).font = Font(color=self.RGB_dict[style])
                # else:
                #     print("no style or style error: %s" % style)
                self.workbook.save(self.test_case_file)
                # self.workbook.save(test_case_file_path)                              # wrong
                info("write cell [%s, %s] successfully" % (row_no, col_no))
            except Exception as e:
                info("write cell [%s, %s] error" % (row_no, col_no))
                raise e
            finally:
                self.workbook.save(self.test_case_file)
        else:
            raise Exception("write cell：insufficient coordinates of cell")

    def write_cell_current_time(self, sheet_obj, row_no = None, col_no = None, style = None):
        current_time = time.strftime("%Y-%m-%d %H:%M:%S")
        # print(current_time)
        if row_no and col_no:
            try:
                sheet_obj.cell(row=row_no, column=col_no).value = current_time          # no raise error, col_no wrong
                if style and style in self.RGB_dict:
                    sheet_obj.cell(row=row_no, column=col_no).font = Font(color=self.RGB_dict[style])
                self.workbook.save(self.test_case_file)
                info("write cell [%s, %s] current time successfully" % (row_no, col_no))
            except Exception as e:
                info("write cell [%s, %s] current time error" % (row_no, col_no))
                raise e
            finally:
                self.workbook.save(self.test_case_file)
        else:
            raise Exception("write time：insufficient coordinates of cell")


if __name__ == "__main__":
    ps = ParseExcel()
    ps.load_workbook(test_case_file_path)
    sheet_obj = ps.get_sheet_by_name("测试用例")         # <Worksheet "测试用例">
    print(sheet_obj.iter_rows())

    print(ps.get_cell_value(sheet_obj, 1, 1))            # use ps to call func
    #
    # print(ps.get_cell_value(sheet_obj, 1))
    # print(ps.get_cell_value(sheet_obj, col_no=1))

    print(ps.get_cell_obj(sheet_obj, 1, 1))              # <Cell '测试用例'.A1>

    ps.write_cell(sheet_obj, 10, 1, "got it", "red")
    ps.write_cell(sheet_obj, 10, 2, "got it", "green")
    ps.write_cell(sheet_obj, 10, 3, "got it")
    ps.write_cell(sheet_obj, 10, 4, "got it", "cc")

    ps.write_cell_current_time(sheet_obj, 12, 1)

